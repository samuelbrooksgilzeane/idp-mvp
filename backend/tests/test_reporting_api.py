import sqlite3
from collections.abc import Iterator
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook  # type: ignore[import-untyped]

from idp_app.core.config import Settings
from idp_app.main import create_app

PDF_ONE = b"%PDF-1.7\n1 0 obj<</Type/Catalog>>endobj\n%%EOF"
PDF_TWO = b"%PDF-1.7\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n%%EOF"


@pytest.fixture
def reporting_client(tmp_path: Path) -> Iterator[tuple[TestClient, Path]]:
    local_dir = tmp_path / "local"
    app = create_app(Settings(_env_file=None, local_data_dir=local_dir))
    with TestClient(app) as client:
        yield client, local_dir / "registry.sqlite3"


def test_summary_case_filter_and_workbook_share_the_same_rows(
    reporting_client: tuple[TestClient, Path],
) -> None:
    client, database_path = reporting_client
    first = _upload(client, "invoice-a.pdf", PDF_ONE, "CASE-A")
    second = _upload(client, "invoice-b.pdf", PDF_TWO, "CASE-B")

    # Opening reporting first initialises all local projection tables.
    assert client.get("/api/results/invoices").json() == []
    _seed_report_rows(database_path, first, second)

    response = client.get("/api/results/invoices")
    assert response.status_code == 200
    assert [row["invoice_number"] for row in response.json()] == ["INV-A", "INV-B"]
    assert response.json()[0] == {
        "document_id": first,
        "file_name": "invoice-a.pdf",
        "case_id": "CASE-A",
        "invoice_index": 0,
        "invoice_number": "INV-A",
        "invoice_date": "2026-08-28",
        "seller_name": "Seller A",
        "currency": "GBP",
        "line_item_count": 2,
        "line_items_sum": "80",
        "total_amount": "75",
        "reconciliation_delta": "0",
        "document_status": "VALIDATED_PASS",
    }
    assert "source_path" not in response.text
    assert "extraction_run_id" not in response.text

    filtered = client.get("/api/results/invoices", params={"case_id": "CASE-B"})
    assert [row["document_id"] for row in filtered.json()] == [second]
    assert filtered.json()[0]["reconciliation_delta"] == "-2"

    cases = client.get("/api/documents/cases")
    assert cases.json() == ["CASE-A", "CASE-B"]
    documents = client.get("/api/documents", params={"case_id": "CASE-A"})
    assert [row["document_id"] for row in documents.json()] == [first]

    exported = client.get("/api/exports/invoices.xlsx", params={"case_id": "CASE-A"})
    assert exported.status_code == 200
    assert exported.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "invoice-results-CASE-A.xlsx" in exported.headers["content-disposition"]
    assert exported.headers["cache-control"] == "no-store"
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    assert workbook.sheetnames == ["Summary", "Line items"]
    summary = list(workbook["Summary"].values)
    lines = list(workbook["Line items"].values)
    assert summary[0][-2:] == ("Reconciliation delta", "Validation outcome")
    assert len(summary) == 2
    assert summary[1][0:5] == (first, "invoice-a.pdf", "CASE-A", 1, "INV-A")
    assert summary[1][-2:] == (0, "VALIDATED_PASS")
    assert len(lines) == 3
    assert {row[0] for row in lines[1:]} == {first}
    assert {row[2] for row in lines[1:]} == {"INV-A"}

    unicode_case = client.get(
        "/api/exports/invoices.xlsx", params={"case_id": "客户 1"}
    )
    assert unicode_case.status_code == 200
    assert unicode_case.headers["content-disposition"].isascii()
    assert workbook["Summary"].freeze_panes == "A2"
    assert workbook["Summary"].sheet_view.showGridLines is False
    assert workbook["Summary"]["J2"].number_format == '#,##0.00;[Red]-#,##0.00'
    assert workbook["Line items"]["F2"].number_format == "#,##0.0000"

    # The verified calculation never assumes a missing signed term is zero.
    with sqlite3.connect(database_path) as connection:
        connection.execute(
            "UPDATE invoice_candidates SET discount_amount = NULL "
            "WHERE extraction_run_id = 'run-a'"
        )
    missing_discount = client.get(
        "/api/results/invoices", params={"case_id": "CASE-A"}
    )
    assert missing_discount.json()[0]["reconciliation_delta"] is None


def test_reported_delta_uses_the_same_signed_terms_as_the_registered_rule(
    reporting_client: tuple[TestClient, Path],
) -> None:
    """A reported delta must never contradict the validation outcome beside it.

    `line_items_reconcile_to_total` reconciles the billed lines *and their stated line
    tax* against the total, so a summary that omitted line tax would report a large
    exception on an invoice the validator passed.
    """
    client, database_path = reporting_client
    first = _upload(client, "invoice-a.pdf", PDF_ONE, "CASE-A")
    second = _upload(client, "invoice-b.pdf", PDF_TWO, "CASE-B")
    assert client.get("/api/results/invoices").json() == []
    _seed_report_rows(database_path, first, second)

    # Lines of 50 + 30 carrying 10 + 6 of stated line tax, less a discount of 5,
    # reconcile exactly to a stated total of 91.
    with sqlite3.connect(database_path) as connection:
        connection.execute(
            "UPDATE invoice_line_candidates SET tax = '10' "
            "WHERE extraction_run_id = 'run-a' AND line_number = 1"
        )
        connection.execute(
            "UPDATE invoice_line_candidates SET tax = '6' "
            "WHERE extraction_run_id = 'run-a' AND line_number = 2"
        )
        connection.execute(
            "UPDATE invoice_candidates SET total_amount = '91' "
            "WHERE extraction_run_id = 'run-a'"
        )
    balanced = client.get("/api/results/invoices", params={"case_id": "CASE-A"})
    assert balanced.json()[0]["reconciliation_delta"] == "0"

    # An unstated line tax is missing, not zero, so the delta is unknown rather than wrong.
    with sqlite3.connect(database_path) as connection:
        connection.execute(
            "UPDATE invoice_line_candidates SET tax = NULL WHERE extraction_run_id = 'run-a'"
        )
    unstated_tax = client.get("/api/results/invoices", params={"case_id": "CASE-A"})
    assert unstated_tax.json()[0]["reconciliation_delta"] is None


def _upload(client: TestClient, filename: str, content: bytes, case_id: str) -> str:
    response = client.post(
        "/api/documents",
        files={"files": (filename, content, "application/pdf")},
        data={"case_id": case_id},
    )
    assert response.status_code == 201
    return str(response.json()["documents"][0]["document_id"])


def _seed_report_rows(database_path: Path, first: str, second: str) -> None:
    with sqlite3.connect(database_path) as connection:
        connection.executemany(
            "INSERT INTO extraction_runs VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("run-a-old", first, "parse-a-old", "invoice", 1, "hash", "2.1", "{}", "{}",
                 None, "EXTRACTED", "test", 0, "2026-08-26T09:00:00+00:00",
                 "2026-08-26T09:01:00+00:00"),
                ("run-a", first, "parse-a", "invoice", 1, "hash", "2.1", "{}", "{}", None,
                 "EXTRACTED", "test", 1, "2026-08-28T09:00:00+00:00",
                 "2026-08-28T09:01:00+00:00"),
                ("run-b", second, "parse-b", "invoice", 1, "hash", "2.1", "{}", "{}", None,
                 "EXTRACTED", "test", 2, "2026-08-27T09:00:00+00:00",
                 "2026-08-27T09:01:00+00:00"),
            ],
        )
        connection.executemany(
            "INSERT INTO invoice_candidates "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("CASE-A", first, "/Volumes/private/old.pdf", "invoice_v1", "INV-OLD",
                 "2026-08-26", "Seller A", "999", "0", "0", "999", "GBP", "run-a-old", 1, 0),
                ("CASE-A", first, "/Volumes/private/a.pdf", "invoice_v1", "INV-A",
                 "2026-08-28", "Seller A", "70", "5", "10", "75", "GBP", "run-a", 1, 0),
                ("CASE-B", second, "/Volumes/private/b.pdf", "invoice_v1", "INV-B",
                 "2026-08-27", "Seller B", "10", "0", "2", "12", "GBP", "run-b", 1, 0),
            ],
        )
        connection.executemany(
            "INSERT INTO invoice_line_candidates VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("run-a-old", first, 0, 1, "Superseded", "1", "999", "0", "999"),
                ("run-a", first, 0, 1, "Advisory", "1", "50", "0", "50"),
                ("run-a", first, 0, 2, "Research", "2", "15", "0", "30"),
                ("run-b", second, 0, 1, "Support", "1", "10", "0", "10"),
            ],
        )
        connection.executemany(
            "INSERT INTO validation_runs VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("validation-a-old", first, "run-a", "invoice", 1, "hash", "1.0",
                 "COMPLETED", "REVIEW_REQUIRED", "test", "2026-08-28T09:01:00+00:00",
                 "2026-08-28T09:02:00+00:00"),
                ("validation-a", first, "run-a", "invoice", 1, "hash", "1.0", "COMPLETED",
                 "VALIDATED_PASS", "test", "2026-08-28T09:02:00+00:00",
                 "2026-08-28T09:03:00+00:00"),
            ],
        )


def test_a_document_stating_several_invoices_reports_one_row_each(
    reporting_client: tuple[TestClient, Path],
) -> None:
    """Each invoice reports its own lines, never the whole document's.

    Joining lines on the extraction run alone would give every invoice in a document every
    other invoice's lines, inflating each sum without ever raising an error.
    """
    client, database_path = reporting_client
    first = _upload(client, "invoice-a.pdf", PDF_ONE, "CASE-A")
    second = _upload(client, "three-invoices.pdf", PDF_TWO, "CASE-MULTI")
    assert client.get("/api/results/invoices").json() == []
    _seed_report_rows(database_path, first, second)

    # run-b states three invoices of 10, 200 and 3000, with one, two and one lines.
    with sqlite3.connect(database_path) as connection:
        connection.executemany(
            "INSERT INTO invoice_candidates "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("CASE-MULTI", second, "/Volumes/private/b.pdf", "invoice_v4", "INV-B2",
                 "2026-08-27", "Seller B", "200", "0", "0", "200", "GBP", "run-b", 4, 1),
                ("CASE-MULTI", second, "/Volumes/private/b.pdf", "invoice_v4", "INV-B3",
                 "2026-08-27", "Seller C", "3000", "0", "0", "3000", "GBP", "run-b", 4, 2),
            ],
        )
        connection.executemany(
            "INSERT INTO invoice_line_candidates VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("run-b", second, 1, 1, "Second, line one", "1", "150", "0", "150"),
                ("run-b", second, 1, 2, "Second, line two", "1", "50", "0", "50"),
                ("run-b", second, 2, 1, "Third", "1", "3000", "0", "3000"),
            ],
        )

    rows = client.get("/api/results/invoices", params={"case_id": "CASE-MULTI"}).json()
    assert [(row["invoice_index"], row["invoice_number"]) for row in rows] == [
        (0, "INV-B"), (1, "INV-B2"), (2, "INV-B3")
    ]
    # Each invoice counts and sums only its own lines.
    assert [row["line_item_count"] for row in rows] == [1, 2, 1]
    assert [row["line_items_sum"] for row in rows] == ["10", "200", "3000"]
    assert [row["reconciliation_delta"] for row in rows] == ["-2", "0", "0"]

    exported = client.get("/api/exports/invoices.xlsx", params={"case_id": "CASE-MULTI"})
    summary = list(load_workbook(BytesIO(exported.content), data_only=True)["Summary"].values)
    lines = list(load_workbook(BytesIO(exported.content), data_only=True)["Line items"].values)
    assert [row[3] for row in summary[1:]] == [1, 2, 3]
    # Every exported line names the invoice it belongs to, so the sheets join on it.
    assert [(row[1], row[2], row[3]) for row in lines[1:]] == [
        (1, "INV-B", 1), (2, "INV-B2", 1), (2, "INV-B2", 2), (3, "INV-B3", 1),
    ]
