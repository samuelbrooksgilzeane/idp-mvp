from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]
from starlette.concurrency import run_in_threadpool

from idp_app.services.document_registry import DatabricksDocumentRegistry
from idp_app.services.documents import DocumentServiceError


@dataclass(frozen=True)
class InvoiceSummaryRecord:
    document_id: str
    file_name: str
    case_id: str | None
    invoice_number: str | None
    invoice_date: date | None
    seller_name: str | None
    currency: str | None
    line_item_count: int
    line_items_sum: Decimal | None
    total_amount: Decimal | None
    reconciliation_delta: Decimal | None
    document_status: str | None


@dataclass(frozen=True)
class InvoiceLineExportRecord:
    document_id: str
    invoice_number: str | None
    line_number: int
    description: str | None
    quantity: Decimal | None
    unit_price: Decimal | None
    tax: Decimal | None
    amount: Decimal | None


class ReportingRepository(Protocol):
    def list_invoice_summaries(
        self, case_id: str | None = None
    ) -> list[InvoiceSummaryRecord]: ...

    def list_invoice_lines(
        self, case_id: str | None = None
    ) -> list[InvoiceLineExportRecord]: ...


class SQLiteReportingRepository:
    def __init__(self, database_path: Path) -> None:
        self._database_path = database_path

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self._database_path)
        connection.row_factory = sqlite3.Row
        return connection

    def list_invoice_summaries(
        self, case_id: str | None = None
    ) -> list[InvoiceSummaryRecord]:
        where = "WHERE documents.case_id = ?" if case_id is not None else ""
        parameters: tuple[str, ...] = (case_id,) if case_id is not None else ()
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                WITH latest_extractions AS (
                  SELECT extraction_run_id, document_id
                  FROM (
                    SELECT extraction_run_id, document_id,
                           ROW_NUMBER() OVER (
                             PARTITION BY document_id
                             ORDER BY completed_at DESC, extraction_run_id DESC
                           ) AS position
                    FROM extraction_runs
                    WHERE status = 'EXTRACTED'
                  )
                  WHERE position = 1
                ),
                latest_validations AS (
                  SELECT extraction_run_id, document_status
                  FROM (
                    SELECT extraction_run_id, document_status,
                           ROW_NUMBER() OVER (
                             PARTITION BY extraction_run_id
                             ORDER BY completed_at DESC, validation_run_id DESC
                           ) AS position
                    FROM validation_runs
                    WHERE status = 'COMPLETED'
                  )
                  WHERE position = 1
                )
                SELECT documents.document_id, documents.file_name, documents.case_id,
                       candidates.invoice_number, candidates.invoice_date,
                       candidates.seller_name, candidates.currency,
                       candidates.discount_amount, candidates.total_amount,
                       latest_validations.document_status,
                       latest_extractions.extraction_run_id
                FROM documents
                JOIN latest_extractions
                  ON latest_extractions.document_id = documents.document_id
                JOIN invoice_candidates AS candidates
                  ON candidates.extraction_run_id = latest_extractions.extraction_run_id
                LEFT JOIN latest_validations
                  ON latest_validations.extraction_run_id = latest_extractions.extraction_run_id
                {where}
                ORDER BY CASE WHEN candidates.invoice_date IS NULL THEN 1 ELSE 0 END,
                         candidates.invoice_date DESC, documents.file_name, documents.document_id
                """,
                parameters,
            ).fetchall()
            run_ids = [str(row["extraction_run_id"]) for row in rows]
            line_terms = _sqlite_line_terms(connection, run_ids)

        summaries: list[InvoiceSummaryRecord] = []
        for row in rows:
            terms = line_terms.get(str(row["extraction_run_id"]), [])
            line_sum = _stated_sum([amount for amount, _ in terms])
            line_tax_sum = _stated_sum([tax for _, tax in terms])
            total = _decimal(row["total_amount"])
            discount = _decimal(row["discount_amount"])
            # The signed terms of the registered line_items_reconcile_to_total rule.
            delta = (
                line_sum + line_tax_sum - discount - total
                if line_sum is not None
                and line_tax_sum is not None
                and discount is not None
                and total is not None
                else None
            )
            summaries.append(
                InvoiceSummaryRecord(
                    document_id=str(row["document_id"]),
                    file_name=str(row["file_name"]),
                    case_id=_optional_text(row["case_id"]),
                    invoice_number=_optional_text(row["invoice_number"]),
                    invoice_date=_date(row["invoice_date"]),
                    seller_name=_optional_text(row["seller_name"]),
                    currency=_optional_text(row["currency"]),
                    line_item_count=len(terms),
                    line_items_sum=line_sum,
                    total_amount=total,
                    reconciliation_delta=delta,
                    document_status=_optional_text(row["document_status"]),
                )
            )
        return summaries

    def list_invoice_lines(
        self, case_id: str | None = None
    ) -> list[InvoiceLineExportRecord]:
        where = "AND documents.case_id = ?" if case_id is not None else ""
        parameters: tuple[str, ...] = (case_id,) if case_id is not None else ()
        with self._connect() as connection:
            rows = connection.execute(
                f"""
                WITH latest_extractions AS (
                  SELECT extraction_run_id, document_id
                  FROM (
                    SELECT extraction_run_id, document_id,
                           ROW_NUMBER() OVER (
                             PARTITION BY document_id
                             ORDER BY completed_at DESC, extraction_run_id DESC
                           ) AS position
                    FROM extraction_runs
                    WHERE status = 'EXTRACTED'
                  )
                  WHERE position = 1
                )
                SELECT documents.document_id, candidates.invoice_number,
                       lines.line_number, lines.description, lines.quantity,
                       lines.unit_price, lines.tax, lines.amount
                FROM latest_extractions
                JOIN documents ON documents.document_id = latest_extractions.document_id
                JOIN invoice_candidates AS candidates
                  ON candidates.extraction_run_id = latest_extractions.extraction_run_id
                JOIN invoice_line_candidates AS lines
                  ON lines.extraction_run_id = latest_extractions.extraction_run_id
                WHERE 1 = 1 {where}
                ORDER BY documents.document_id, lines.line_number
                """,
                parameters,
            ).fetchall()
        return [_line_from_values(row) for row in rows]


class DatabricksReportingRepository:
    def __init__(
        self,
        sql_client: DatabricksDocumentRegistry,
        catalog: str,
        project_schema: str,
        table_prefix: str,
    ) -> None:
        self._sql = sql_client
        prefix = f"{catalog}.{project_schema}.{table_prefix}"
        self._summary = f"{prefix}_invoice_summary"
        self._lines = f"{prefix}_invoice_line_candidates"

    def list_invoice_summaries(
        self, case_id: str | None = None
    ) -> list[InvoiceSummaryRecord]:
        where = " WHERE case_id = :case_id" if case_id is not None else ""
        rows = self._sql.execute_sql(
            "SELECT document_id, file_name, case_id, invoice_number, invoice_date, "
            "seller_name, currency, line_item_count, line_items_sum, total_amount, "
            f"reconciliation_delta, document_status FROM {self._summary}{where} "
            "ORDER BY invoice_date DESC NULLS LAST, file_name, document_id LIMIT 500",
            {"case_id": case_id} if case_id is not None else None,
        )
        return [_summary_from_values(row) for row in rows]

    def list_invoice_lines(
        self, case_id: str | None = None
    ) -> list[InvoiceLineExportRecord]:
        where = " WHERE summary.case_id = :case_id" if case_id is not None else ""
        rows = self._sql.execute_sql(
            "SELECT summary.document_id, summary.invoice_number, lines.line_number, "
            "lines.description, lines.quantity, lines.unit_price, lines.tax, lines.amount "
            f"FROM {self._summary} AS summary JOIN {self._lines} AS lines "
            "ON lines.extraction_run_id = summary.extraction_run_id"
            f"{where} ORDER BY summary.document_id, lines.line_number LIMIT 50000",
            {"case_id": case_id} if case_id is not None else None,
        )
        return [_line_from_values(row) for row in rows]


class ReportingService:
    def __init__(self, repository: ReportingRepository) -> None:
        self._repository = repository

    async def list_invoice_summaries(
        self, case_id: str | None = None
    ) -> list[InvoiceSummaryRecord]:
        try:
            return await run_in_threadpool(self._repository.list_invoice_summaries, case_id)
        except Exception as error:
            raise DocumentServiceError(
                "REPORT_READ_FAILED", "Invoice results could not be loaded.", 502
            ) from error

    async def export_invoice_workbook(self, case_id: str | None = None) -> BytesIO:
        try:
            summaries, lines = await run_in_threadpool(
                _load_report_rows, self._repository, case_id
            )
            return await run_in_threadpool(_build_workbook, summaries, lines, case_id)
        except Exception as error:
            raise DocumentServiceError(
                "EXPORT_FAILED", "The invoice workbook could not be generated.", 502
            ) from error


def _load_report_rows(
    repository: ReportingRepository, case_id: str | None
) -> tuple[list[InvoiceSummaryRecord], list[InvoiceLineExportRecord]]:
    return (
        repository.list_invoice_summaries(case_id),
        repository.list_invoice_lines(case_id),
    )


def _build_workbook(
    summaries: list[InvoiceSummaryRecord],
    lines: list[InvoiceLineExportRecord],
    case_id: str | None,
) -> BytesIO:
    workbook = Workbook()
    workbook.properties.title = "Invoice extraction results"
    workbook.properties.subject = f"Case {case_id}" if case_id else "All cases"
    workbook.properties.creator = "IDP MVP"

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_headers = [
        "Document ID", "File name", "Case ID", "Invoice number", "Invoice date",
        "Seller", "Currency", "Line item count", "Line items sum", "Stated total",
        "Reconciliation delta", "Validation outcome",
    ]
    summary_sheet.append(summary_headers)
    for summary_row in summaries:
        summary_sheet.append([
            summary_row.document_id, summary_row.file_name, summary_row.case_id,
            summary_row.invoice_number, summary_row.invoice_date, summary_row.seller_name,
            summary_row.currency, summary_row.line_item_count,
            _excel_decimal(summary_row.line_items_sum),
            _excel_decimal(summary_row.total_amount),
            _excel_decimal(summary_row.reconciliation_delta), summary_row.document_status,
        ])

    lines_sheet = workbook.create_sheet("Line items")
    line_headers = [
        "Document ID", "Invoice number", "Line number", "Description", "Quantity",
        "Unit price", "Tax", "Amount",
    ]
    lines_sheet.append(line_headers)
    for line_row in lines:
        lines_sheet.append([
            line_row.document_id, line_row.invoice_number, line_row.line_number,
            line_row.description, _excel_decimal(line_row.quantity),
            _excel_decimal(line_row.unit_price), _excel_decimal(line_row.tax),
            _excel_decimal(line_row.amount),
        ])

    _format_sheet(summary_sheet, len(summaries), (9, 10, 11), (5,))
    _format_sheet(lines_sheet, len(lines), (6, 7, 8), ())
    for cells in lines_sheet.iter_cols(min_col=5, max_col=5, min_row=2):
        for cell in cells:
            cell.number_format = "#,##0.0000"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_sheet(
    sheet: Any, data_rows: int, currency_columns: tuple[int, ...], date_columns: tuple[int, ...]
) -> None:
    header_fill = PatternFill("solid", fgColor="202420")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if data_rows:
        table = Table(displayName=f"{sheet.title.replace(' ', '')}Table", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
    for column in currency_columns:
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00;[Red]-#,##0.00'
    for column in date_columns:
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "yyyy-mm-dd"
    for cells in sheet.columns:
        letter = cells[0].column_letter
        maximum = max((len(str(cell.value)) if cell.value is not None else 0) for cell in cells)
        sheet.column_dimensions[letter].width = min(max(maximum + 2, 12), 38)


def _sqlite_line_terms(
    connection: sqlite3.Connection, run_ids: list[str]
) -> dict[str, list[tuple[Decimal | None, Decimal | None]]]:
    """Return the stated amount and stated tax of every billed line, per extraction run."""
    if not run_ids:
        return {}
    placeholders = ", ".join("?" for _ in run_ids)
    rows = connection.execute(
        f"SELECT extraction_run_id, amount, tax FROM invoice_line_candidates "
        f"WHERE extraction_run_id IN ({placeholders}) ORDER BY extraction_run_id, line_number",
        run_ids,
    ).fetchall()
    result: dict[str, list[tuple[Decimal | None, Decimal | None]]] = {
        run_id: [] for run_id in run_ids
    }
    for row in rows:
        result[str(row["extraction_run_id"])].append(
            (_decimal(row["amount"]), _decimal(row["tax"]))
        )
    return result


def _stated_sum(values: list[Decimal | None]) -> Decimal | None:
    """Sum the stated values, or return None when the source stated none of them.

    This mirrors the validator: an aggregate over zero stated instances is missing, not
    zero, so it can never let a reconciliation appear to balance.
    """
    stated = [value for value in values if value is not None]
    return sum(stated, Decimal(0)) if stated else None


def _summary_from_values(values: Any) -> InvoiceSummaryRecord:
    return InvoiceSummaryRecord(
        document_id=str(values[0]), file_name=str(values[1]),
        case_id=_optional_text(values[2]), invoice_number=_optional_text(values[3]),
        invoice_date=_date(values[4]), seller_name=_optional_text(values[5]),
        currency=_optional_text(values[6]), line_item_count=int(values[7]),
        line_items_sum=_decimal(values[8]), total_amount=_decimal(values[9]),
        reconciliation_delta=_decimal(values[10]), document_status=_optional_text(values[11]),
    )


def _line_from_values(values: Any) -> InvoiceLineExportRecord:
    return InvoiceLineExportRecord(
        document_id=str(values[0]), invoice_number=_optional_text(values[1]),
        line_number=int(values[2]), description=_optional_text(values[3]),
        quantity=_decimal(values[4]), unit_price=_decimal(values[5]),
        tax=_decimal(values[6]), amount=_decimal(values[7]),
    )


def _optional_text(value: Any) -> str | None:
    return str(value) if value is not None else None


def _decimal(value: Any) -> Decimal | None:
    return Decimal(str(value)) if value is not None else None


def _date(value: Any) -> date | None:
    if value is None or isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _excel_decimal(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None
