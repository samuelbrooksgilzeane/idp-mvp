"""Schema-driven exports: section 7 of the generalized IDP plan.

Replaces the invoice-only `/api/exports/invoices.xlsx` with a generic `POST /api/exports` that
works for any schema shape:

  * A flat schema exports as one worksheet (or one CSV).
  * A singleton nested object flattens into its containing table using dotted column names.
  * Every *repeated* collection (an array, at any depth) becomes its own related table/sheet,
    related back to its parent table by `_parent_record_id`.
  * Sibling arrays are never cross-joined: each array's rows live in their own table.

This never constructs a SQL identifier from a user-supplied field name: sheet and column names
are derived strings used only as spreadsheet labels, and the underlying data always comes from
the already-validated, already-retained `ai_extract` result via `walk_extraction`.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]

from idp_app.services.document_models import (
    ExtractedRecordRow,
    ExtractionRunRecord,
    GenericFieldRow,
)
from idp_app.services.schema_models import SchemaRecord, schema_leaves

_SANITIZE = re.compile(r"[^A-Za-z0-9]+")


@dataclass
class ExportTable:
    name: str
    schema_path: str
    columns: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)


def _table_name(schema_path: str) -> str:
    if schema_path == "$":
        return "Document"
    last = re.split(r"\.(?![^\[]*\])", schema_path.rstrip("[]"))[-1]
    words = [part for part in _SANITIZE.split(last) if part]
    return "_".join(word.capitalize() for word in words) or "Records"


def build_export_tables(
    run: ExtractionRunRecord,
    records: list[ExtractedRecordRow],
    fields: list[GenericFieldRow],
    document_name: str,
) -> list[ExportTable]:
    """Project one run's recursive walk into relational, spreadsheet-ready tables.

    A "table record" is the root record or any repeated array item -- these become rows in a
    named table. Any other record (a singleton nested object) is a "virtual" record: its
    scalar fields are flattened onto its nearest table-record ancestor's row using a dotted
    column name, rather than becoming a table of their own.
    """
    by_id = {record.record_id: record for record in records}

    def _owning_table_id(record_id: str | None) -> str | None:
        while record_id is not None:
            record = by_id.get(record_id)
            if record is None:
                return None
            if record.parent_record_id is None or record.ordinal is not None:
                return record.record_id
            record_id = record.parent_record_id
        return None

    def _dotted_prefix(record_id: str) -> str:
        """The dotted path from a record up to (but excluding) its owning table row."""
        segments: list[str] = []
        current = by_id.get(record_id)
        while (
            current is not None
            and current.parent_record_id is not None
            and current.ordinal is None
        ):
            name = current.schema_path.rsplit(".", 1)[-1].rstrip("[]")
            segments.append(name)
            parent = by_id.get(current.parent_record_id)
            if parent is None or parent.ordinal is not None or parent.parent_record_id is None:
                break
            current = parent
        return ".".join(reversed(segments))

    tables: dict[str, ExportTable] = {}
    table_row_by_record: dict[str, dict[str, Any]] = {}

    table_records = [r for r in records if r.parent_record_id is None or r.ordinal is not None]
    for record in table_records:
        table = tables.setdefault(
            record.schema_path,
            ExportTable(name=_table_name(record.schema_path), schema_path=record.schema_path),
        )
        parent_table_id = (
            _owning_table_id(record.parent_record_id) if record.parent_record_id else None
        )
        row: dict[str, Any] = {
            "_document_id": run.document_id,
            "_record_id": record.record_id,
            "_parent_record_id": parent_table_id,
            "_ordinal": record.ordinal,
            "_document_name": document_name,
            "_extraction_run_id": run.extraction_run_id,
            "_schema_version": run.schema_version,
            "_extracted_at": (run.completed_at or run.started_at).isoformat(),
        }
        table.rows.append(row)
        table_row_by_record[record.record_id] = row

    for entry in fields:
        table_record_id = _owning_table_id(entry.record_id)
        if table_record_id is None:
            continue
        table_row = table_row_by_record.get(table_record_id)
        if table_row is None:
            continue
        prefix = _dotted_prefix(entry.record_id)
        column = f"{prefix}.{entry.field_name}" if prefix else entry.field_name
        table_row[column] = entry.value

    for table in tables.values():
        seen: list[str] = []
        for row in table.rows:
            for key in row:
                if key not in seen:
                    seen.append(key)
        relationship = [
            "_document_id",
            "_record_id",
            "_parent_record_id",
            "_ordinal",
            "_document_name",
            "_extraction_run_id",
            "_schema_version",
            "_extracted_at",
        ]
        table.columns = relationship + [key for key in seen if key not in relationship]

    return list(tables.values())


def data_dictionary_rows(schema: SchemaRecord) -> list[tuple[str, str, str, str]]:
    return [
        (path, path.rsplit(".", 1)[-1].rstrip("[]"), definition.type, definition.description)
        for path, definition in schema_leaves(schema.ai_extract_schema)
    ]


def build_workbook(
    tables_by_run: list[tuple[ExtractionRunRecord, SchemaRecord, list[ExportTable]]],
) -> io.BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "Extraction results"
    workbook.properties.creator = "IDP MVP"

    merged: dict[str, ExportTable] = {}
    for _run, _schema, tables in tables_by_run:
        for table in tables:
            existing = merged.get(table.name)
            if existing is None:
                merged[table.name] = ExportTable(
                    name=table.name, schema_path=table.schema_path,
                    columns=list(table.columns), rows=list(table.rows),
                )
            else:
                for column in table.columns:
                    if column not in existing.columns:
                        existing.columns.append(column)
                existing.rows.extend(table.rows)

    for table in merged.values():
        sheet = workbook.create_sheet(table.name[:31])
        sheet.append(table.columns)
        for row in table.rows:
            sheet.append([_excel_value(row.get(column)) for column in table.columns])
        _format_sheet(sheet, len(table.rows))

    dictionary = workbook.create_sheet("Data_dictionary")
    dictionary.append(["Schema path", "Field name", "Type", "Description"])
    seen_paths: set[tuple[str, str]] = set()
    for _run, schema, _tables in tables_by_run:
        for dictionary_row in data_dictionary_rows(schema):
            key = (schema.schema_id, dictionary_row[0])
            if key in seen_paths:
                continue
            seen_paths.add(key)
            dictionary.append(list(dictionary_row))
    _format_sheet(dictionary, len(seen_paths))

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_csv_bundle(
    tables_by_run: list[tuple[ExtractionRunRecord, SchemaRecord, list[ExportTable]]],
) -> io.BytesIO:
    """One CSV per collection, bundled into a ZIP -- the flat-file equivalent of the workbook."""
    merged: dict[str, ExportTable] = {}
    for _run, _schema, tables in tables_by_run:
        for table in tables:
            existing = merged.get(table.name)
            if existing is None:
                merged[table.name] = ExportTable(
                    name=table.name, schema_path=table.schema_path,
                    columns=list(table.columns), rows=list(table.rows),
                )
            else:
                for column in table.columns:
                    if column not in existing.columns:
                        existing.columns.append(column)
                existing.rows.extend(table.rows)

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for table in merged.values():
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(table.columns)
            for row in table.rows:
                writer.writerow([_csv_value(row.get(column)) for column in table.columns])
            archive.writestr(f"{table.name}.csv", buffer.getvalue())

        dictionary_buffer = io.StringIO()
        writer = csv.writer(dictionary_buffer)
        writer.writerow(["Schema path", "Field name", "Type", "Description"])
        seen_paths: set[tuple[str, str]] = set()
        for _run, schema, _tables in tables_by_run:
            for dictionary_row in data_dictionary_rows(schema):
                key = (schema.schema_id, dictionary_row[0])
                if key in seen_paths:
                    continue
                seen_paths.add(key)
                writer.writerow(dictionary_row)
        archive.writestr("Data_dictionary.csv", dictionary_buffer.getvalue())

    output.seek(0)
    return output


def _excel_value(value: Any) -> Any:
    if isinstance(value, dict | list):
        import json

        return json.dumps(value, ensure_ascii=False)
    return value


def _csv_value(value: Any) -> Any:
    return _excel_value(value)


def _format_sheet(sheet: Any, data_rows: int) -> None:
    header_fill = PatternFill("solid", fgColor="202420")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    if sheet.dimensions and sheet.dimensions != "A1:A1":
        sheet.auto_filter.ref = sheet.dimensions
    for cells in sheet.columns:
        letter = cells[0].column_letter
        maximum = max((len(str(cell.value)) if cell.value is not None else 0) for cell in cells)
        sheet.column_dimensions[letter].width = min(max(maximum + 2, 12), 38)
